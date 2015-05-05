import pandas
import random
from openpyxl import load_workbook
import time

start_time = time.time()

###----------CONSTANTS------------###
template = 'Output Template 2.8.4V.xlsx'
output = 'Output Template 2.8.4V Filled.xlsx'
XLSXengine = 'openpyxl'
output_sheets = ["Data-Front End", "Data-Reactor", "Data-Recycle", "Data-Inventories", "Data-Economics", "Data-Wildcard"]
metrics = {
"U Resources Mined": [5,2, "Data-Front End"],
}


###----------FUNCTIONS------------###

#initialize the excel write, returns the write object
def init(tpl = template, out = output, eng = XLSXengine, sheet = output_sheets):
    book = load_workbook(tpl)
    writer = pandas.ExcelWriter(out, engine = eng)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    return writer

#exports the excel sheet
def export(writer_object):
    writer_object.save()

#writes dataframe to excel spreadsheet    
def toExcel(dFrame, writer_object, worksheet, col_to_write, row, col):
    dFrame.to_excel(writer_object, worksheet, columns = col_to_write, header = False, index = False, index_label = None, merge_cells = False, engine = XLSXengine, startrow = row, startcol = col)

#writes metric
def writeMetric(metric, dFrame):
    row = metrics[metric][0]
    col = metrics[metric][1]
    sheet = metrics[metric][2]
    toExcel(dFrame, sheet, [0,1,2,3], row, col)
    
#creates a random matrix for testing purposes
def matrix(h, w):
    hor = []
    ver = []
    for j in range(0, h):
        for i in range(0, w):
            ver.append(random.randrange(0,h))
        hor.append(ver)
        ver = []
    return hor
m = matrix(200, 4)
#print(m)
df = pandas.DataFrame(m)
#print(df)




#book = load_workbook('Output Template 2.8.4V.xlsx')
#writer = pandas.ExcelWriter('Output Template 2.8.4V Filled.xlsx', engine ='openpyxl') 
#writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#print(writer.sheets)
#df.to_excel(writer, "Data-Front End", columns=[0,1,2,3], header = False, index = False, index_label=None, merge_cells=False, engine = 'openpyxl', startrow = 5, startcol = 2)
#this will write starting at the top left corner.
#writer.save()

elapsed_time = time.time() - start_time
print(elapsed_time)


##todo: write a python function convert a and aa and ab to their respective index numbers
##write forloop to write multiple dataframes to exel